#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 Excel 文件导入产品数据到 products.json
使用方法：python import_from_excel.py

Excel 列结构:
1: 产品名称
2: 产品代码
3: 产品简称
4: 挂钩标的
5: 产品提前终止日
6: 期初观察日
7: 期末观察日
8:产品成立日
9: 雪球产品类型
10: 雪球结构期限
11: 敲出价格系数 (简写) - 文本描述
12: 敲入价格系数 (简写) - 百分比
13: 派息价格
14: 敲出业绩报酬计提基准 Bi（年化）
15: 敲出事件情形
16: 未敲入未敲出事件情形
17: 敲入事件情形
"""

import json
import os
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook

# 导入交易日历
from trading_calendar import is_trading_day

# 文件路径
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(os.path.expanduser('~'), 'Downloads')
OUTPUT_DIR = os.path.join(BASE_DIR, 'data')

# 查找 Excel 文件
EXCEL_FILE = os.path.join(UPLOAD_DIR, '华夏资本产品列表-2026.08.16.xlsx')
if not os.path.exists(EXCEL_FILE):
    print(f"错误：未找到 Excel 文件 {EXCEL_FILE}")
    exit(1)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'products.json')

print(f"找到 Excel 文件：{EXCEL_FILE}")

def parse_product_type(type_str):
    """将 Excel 中的产品类型转换为系统标准类型"""
    if not type_str:
        return '欧式'
    
    type_str = str(type_str).strip()
    
    if '降敲' in type_str and '早利' in type_str:
        return '欧式早利'
    if '早利' in type_str:
        return '欧式早利'
    if '凤凰' in type_str:
        return '欧式凤凰'
    if type_str == 'FCN':
        return 'FCN'
    if type_str == '欧式':
        return '欧式'
    
    return type_str

def parse_term(term_val):
    """解析期限字段，支持"36 个月"格式"""
    if not term_val:
        return 36
    
    term_str = str(term_val).strip()
    
    try:
        return int(term_str)
    except:
        pass
    
    match = re.search(r'(\d+)', term_str)
    if match:
        return int(match.group(1))
    
    return 36

def parse_knock_out_schedule(text, term_months):
    """解析敲出价格系数文本，生成逐月系数列表"""
    if not text:
        return []
    
    text = str(text).strip()
    
    # 匹配：从第【3】个月开始，【102%】每月递减【0.5%】，最后一个月为【66%】
    # 或：最后一次敲出价为【66%】
    pattern = r'从第【?(\d+)】?个月开始，【?(\d+(?:\.\d+)?)%】?每月递减【?(\d+(?:\.\d+)?)%】?，(?:最后一个月为|最后一次敲出价为)【?(\d+(?:\.\d+)?)%】?'
    m = re.search(pattern, text)
    
    if not m:
        # 回退：尝试提取单个百分比
        single = re.search(r'(\d+(?:\.\d+)?)\s*%', text)
        if single:
            ratio = float(single.group(1))
            return [{"month": i, "ratio": ratio} for i in range(3, term_months + 1)]
        return []
    
    start_month = int(m.group(1))
    first_ratio = float(m.group(2))
    step = float(m.group(3))
    last_ratio = float(m.group(4))
    
    schedule = []
    for month in range(start_month, term_months + 1):
        ratio = first_ratio - (month - start_month) * step
        # 强制截断：最后一个月必须等于 last_ratio
        if month == term_months:
            ratio = last_ratio
        elif ratio <= last_ratio:
            ratio = last_ratio
        schedule.append({"month": month, "ratio": round(ratio, 2)})
    
    return schedule

def parse_knock_in_ratio(text):
    """从敲入价格系数文本中提取敲入比例"""
    if not text:
        return 100
    
    text = str(text).strip()
    
    # 尝试直接转换为浮点数
    try:
        return float(text)
    except:
        pass
    
    # 提取百分比
    match = re.search(r'(\d+(?:\.\d+)?)[%\s]', text)
    if match:
        return float(match.group(1))
    
    return 100

def parse_date(date_val):
    """解析 Excel 日期"""
    if not date_val:
        return None
    
    if isinstance(date_val, str):
        date_val = date_val.strip()
        if date_val == '/' or date_val == '':
            return None
        try:
            datetime.strptime(date_val, '%Y-%m-%d')
            return date_val
        except:
            return None
    
    try:
        # Excel 日期序列号
        base = datetime(1899, 12, 30)
        days = int(float(date_val))
        result = base + timedelta(days=days)
        return result.strftime('%Y-%m-%d')
    except:
        return None

def parse_underlying(underlying_raw):
    """解析挂钩标的，提取名称和代码"""
    if not underlying_raw:
        return '', ''
    
    underlying_raw = str(underlying_raw).strip()
    underlying = underlying_raw.split('(')[0].strip()
    underlying_code = ''
    
    if '代码:' in underlying_raw:
        start = underlying_raw.find('代码:') + 3
        end = underlying_raw.find(')', start)
        if end > start:
            underlying_code = underlying_raw[start:end].strip()
    
    return underlying, underlying_code

def calc_observation_calendar(initial_date: str, start_month: int, term_months: int) -> list[dict]:
    """生成敲出观察日历"""
    import calendar as cal
    
    dt = datetime.strptime(initial_date, "%Y-%m-%d")
    observations = []
    
    for m in range(start_month, term_months + 1):
        new_month = dt.month + m
        new_year = dt.year + (new_month - 1) // 12
        new_month = (new_month - 1) % 12 + 1
        _, last_day = cal.monthrange(new_year, new_month)
        
        obs = dt.replace(year=new_year, month=new_month, day=min(dt.day, last_day))
        obs_str = obs.strftime("%Y-%m-%d")
        
        # 遇非交易日顺延
        while not is_trading_day(obs_str):
            obs += timedelta(days=1)
            obs_str = obs.strftime("%Y-%m-%d")
        
        observations.append({
            "period": m,
            "date": obs_str,
            "price": None,
            "status": "待观察"
        })
    
    return observations

def import_products():
    """主导入函数"""
    print(f"正在读取 Excel 文件：{EXCEL_FILE}")
    
    if not os.path.exists(EXCEL_FILE):
        print(f"错误：文件不存在 {EXCEL_FILE}")
        return False
    
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"读取 Excel 失败：{e}")
        return False
    
    products = []
    
    # 从第 2 行开始读取（第 1 行是表头）
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # 解析各列数据
            product_name = str(row[0]).strip() if row[0] else ''
            product_code = str(row[1]).strip() if row[1] else ''
            product_short_name = str(row[2]).strip() if row[2] else ''
            
            # 挂钩标的
            underlying, underlying_code = parse_underlying(row[3])
            
            # 日期字段
            # 列 5: 产品提前终止日 (暂时不用)
            # 列 6: 期初观察日
            # 列 7: 期末观察日
            # 列 8: 产品成立日
            initial_obs_date = parse_date(row[5])
            final_obs_date = parse_date(row[6])
            establish_date = parse_date(row[7])
            
            # 产品类型
            product_type_raw = str(row[8]).strip() if row[8] else '欧式'
            product_type = parse_product_type(product_type_raw)
            
            # 期限
            term_months = parse_term(row[9])
            
            # 敲出敲入系数
            knock_out_schedule = parse_knock_out_schedule(row[10], term_months)
            knock_in_ratio = parse_knock_in_ratio(row[11])
            
            # 派息价格
            dividend_ratio = None
            if row[12] and str(row[12]) not in ['None', '/', '']:
                try:
                    dividend_ratio = float(str(row[12]).replace('%', ''))
                except:
                    pass
            
            # 条件说明
            knock_out_condition = str(row[14]).strip() if row[14] and str(row[14]) != '/' else ''
            knock_in_condition = str(row[16]).strip() if row[16] and str(row[16]) != '/' else ''
            
            # 计算敲入价格和首月敲出价格（用于兼容）
            initial_price = 10000.00  # 基准价格，后续会被 updater.py 覆盖
            knock_in_price = round(initial_price * (knock_in_ratio / 100), 2)
            
            # 首月敲出系数
            knock_out_ratio = knock_out_schedule[0]['ratio'] if knock_out_schedule else 100.0
            
            # 生成观察日历
            observation_history = []
            if initial_obs_date:
                observation_history = calc_observation_calendar(
                    initial_obs_date,
                    3,  # 敲出观察从第3个月开始
                    term_months
                )
            
            # 确定状态
            current_date = datetime.now().strftime('%Y-%m-%d')
            if establish_date and current_date < establish_date:
                status = '待期初'
            else:
                status = '存续中'
            
            # 计算下次观察日（从观察日历取第一个）
            next_obs = None
            days_to_next = 0
            if observation_history:
                next_obs = observation_history[0]['date']
                days_to_next = max(0, (datetime.strptime(next_obs, '%Y-%m-%d') - datetime.now()).days)
            elif initial_obs_date:
                next_obs = initial_obs_date
                days_to_next = max(0, (datetime.strptime(next_obs, '%Y-%m-%d') - datetime.now()).days)
            
            product = {
                'id': product_code,
                'name': product_name,
                'short_name': product_short_name,
                'code': product_code,
                'underlying': underlying,
                'underlying_code': underlying_code,
                'type': product_type,
                'term_months': term_months,
                'initial_observation_date': initial_obs_date,
                'final_observation_date': final_obs_date,
                'establishment_date': establish_date,
                'knock_out_ratio': knock_out_ratio,
                'knock_out_schedule': knock_out_schedule,
                'knock_in_ratio': knock_in_ratio,
                'dividend_ratio': dividend_ratio,
                'knock_out_start_month': 3,
                'knock_out_condition': knock_out_condition,
                'knock_in_condition': knock_in_condition,
                'status': status,
                'current_price': None,
                'initial_price': initial_price,
                'knock_in_price': knock_in_price,
                'next_observation_date': next_obs,
                'days_to_next': max(0, days_to_next),
                'observation_history': observation_history
            }
            
            products.append(product)
            print(f"  ✓ 导入：{product_code} - {product_short_name} ({product_type}, KO: {knock_out_ratio}%, KI: {knock_in_ratio}%)")
            
        except Exception as e:
            print(f"  ⚠ 导入第{row_idx}行失败：{e}")
            continue
    
    # 创建输出数据结构
    output_data = {
        'last_update': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'products': products,
        'config': {
            'password': 'huaxia2026',
            'update_time': '16:00',
            'timezone': 'Asia/Shanghai',
            'display_fields': {
                'list_view': ['name', 'code', 'underlying', 'current_price', 'status', 'type', 
                             'term_months', 'knock_out_ratio', 'knock_in_ratio', 'days_to_next'],
                'detail_view': 'all'
            }
        }
    }
    
    # 确保目录存在
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    
    # 保存到 JSON 文件
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ 导入完成！共导入 {len(products)} 只产品")
    print(f"📁 数据已保存到：{OUTPUT_FILE}")
    print(f"\n下一步：")
    print(f"1. 刷新页面查看导入的产品")
    print(f"2. 在管理后台可以进一步编辑")
    
    return True

if __name__ == '__main__':
    import_products()
